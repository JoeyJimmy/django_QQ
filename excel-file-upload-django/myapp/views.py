from django.shortcuts import render
import openpyxl

import json
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd


error={'box_error':[]}


def index(request):
    if "GET" == request.method:
        return render(request, 'myapp/index.html', {})
    else:
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        sheets = wb.sheetnames
        print(sheets)

        # getting a particular sheet
        worksheet = wb["Sheet1"]
        print(worksheet)

        # getting active sheet
        active_sheet = wb.active
        print(active_sheet)

        # reading a cell
        print(worksheet["A1"].value)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
                print(cell.value)
            excel_data.append(row_data)

        
        error['box_error'].clear()
        context = data_processing_box(excel_data)
        #context.append(data_processing_box(excel_data))
        #context.append(data_processing_pie(excel_data))
        
        
        return render(request, 'myapp/index.html', {'context':json.dumps(context),'error':json.dumps(error)})


def data_processing_box(excel_data):

    #存title
    title = excel_data[0]
    data = excel_data.copy()
    del data[0]
    #轉float
    df = pd.DataFrame(data, columns=title, dtype='float64')
    #設定會傳入做四分位數的title名稱
    title_list = ['HEIGHT', 'WEIGHT', 'PULSE', 'BLP', 'BHP', 'TC', 'TG', 'BUN', 'CREA', 'GOT', 'GPT']
    context=[]
    for num in range(len(title)):
        title[num] = str.upper(title[num])
        if title[num] in title_list:
            try:
                count = np.sum(df[title[num]].values == "None")
                #print(count)
                df[title[num]] = df[title[num]].replace('None',"")
                df[title[num]] = pd.to_numeric(df[title[num]], errors='ignore')
                tmp = {}
                tmp['title'] = title[num]
                tmp['Q'] = []
                tmp['Q'].append(df[title[num]].quantile(0))
                tmp['Q'].append(df[title[num]].quantile(0.25))
                tmp['Q'].append(df[title[num]].quantile(0.5))
                tmp['Q'].append(df[title[num]].quantile(0.75))
                tmp['Q'].append(df[title[num]].quantile(1))
                tmp['sum']= str(len(df[title[num]])-count)
                context.append(tmp)
            except:
                #print(title[num],'None')
                error['box_error'].append(title[num])
    print(context)
    return context




def check_remove(title,key):
    if key in title :
        title.remove(key)








